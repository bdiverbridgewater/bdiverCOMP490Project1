import re
import sqlite3
import sys
from typing import Tuple

import PySide6.QtWidgets
from PySide6.QtQuick import QQuickWindow, QSGRendererInterface
from openpyxl import load_workbook
from serpapi import google_search

import key_secrets
from FirstWindow import FirstWindow


def job_search(result_offset) -> dict:
    params = {
        "api_key": key_secrets.api_key,
        "engine": "google_jobs",
        "google_domain": "google.com",
        "q": "Software Engineer",
        "hl": "en",
        "gl": "us",
        "location": "Boston, Massachusetts, United States",
        "start": result_offset
    }
    search = google_search.GoogleSearch(params)
    results = search.get_dict()
    return results.get("jobs_results")


def get_jobs_data(jobs):
    jobs_data = []
    index = 0
    while index < 10:
        job = jobs[index]
        job_id = job.get("job_id")
        job_title = job.get("title")
        company_name = job.get("company_name")
        location = job.get("location")
        description = job.get("description")
        related_link = job["related_links"][0].get("link")
        work_from_home = job["detected_extensions"].get("work_from_home")
        time_since_posting = job["detected_extensions"].get("posted_at")
        try:
            benefits_section = job.get("job_highlights")[2]
        except IndexError:
            benefits_section = None

        min_salary, max_salary = get_salary(benefits_section, job["description"])
        qualifications = list_to_string(job["job_highlights"][0]["items"])
        job_data = [job_id, job_title, company_name, location, description, related_link, work_from_home,
                    time_since_posting, min_salary, max_salary, qualifications]
        jobs_data.append(job_data)
        index += 1
    return jobs_data


def list_to_string(list_in):
    str_out = ''
    index = 0
    for _ in list_in:
        str_out = str_out + list_in[index]
        index += 1
    return str_out


def try_to_get_salary(job):
    try:
        salary = job["job_highlights"][2]["items"][0]
    except IndexError:
        salary = None
    return salary


def open_database(file_name: str) -> Tuple[sqlite3.Connection, sqlite3.Cursor]:
    database_connection = sqlite3.connect(file_name)
    cursor = database_connection.cursor()
    return database_connection, cursor


def setup_database(cursor: sqlite3.Cursor):
    cursor.execute('''CREATE TABLE IF NOT EXISTS jobs(
        job_id TEXT PRIMARY KEY,
        job_title TEXT,
        company_name TEXT,
        location TEXT,
        description TEXT,
        related_link TEXT,
        work_from_home INTEGER DEFAULT FALSE,
        time_since_posting TEXT,
        min_salary INTEGER,
        max_salary INTEGER
        );''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS qualifications(
        job_id TEXT,
        qualifications TEXT,
        FOREIGN KEY (job_id) REFERENCES jobs(job_id)
        );''')


def close_database(connection: sqlite3.Connection):
    connection.commit()
    connection.close()


def insert_job_to_database(job_data, cursor):
    jobs_table_data = job_data[:10]
    qualifications_data = [job_data[0], job_data[10]]
    cursor.executemany('''INSERT OR IGNORE INTO jobs
    VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?);''', (jobs_table_data,))
    cursor.executemany('''INSERT OR IGNORE INTO qualifications VALUES(?, ?);''', (qualifications_data,))


def insert_jobs_to_database(jobs_data, cursor):
    for job in jobs_data:
        insert_job_to_database(job, cursor)


def get_excel_jobs():
    wb = load_workbook('Sprint3Data.xlsx')
    jobs_sheet = wb['Comp490 Jobs']
    excel_jobs = []
    for row in jobs_sheet.iter_rows(min_row=2, values_only=True):
        company_name = row[0]
        job_id = row[2]
        location = row[4]
        salary_max = row[6]
        salary_min = row[7]
        job_title = row[9]
        time_since_posting = row[1]
        description = None
        related_link = None
        work_from_home = False
        qualifications = None
        job_data = [job_id, job_title, company_name, location, description, related_link, work_from_home,
                    time_since_posting, salary_min, salary_max, qualifications]
        excel_jobs.append(job_data)
    return excel_jobs


def get_salary(benefits_section: dict, job_description: str):
    """this is more complicated than you were required to do, I'm looking in several places for salary info"""
    min_salary = 0
    max_salary = 0
    if benefits_section:  # if we got a dictionary with stuff in it
        for benefit_item in benefits_section['items']:
            if 'range' in benefit_item.lower():
                # from https://stackoverflow.com/questions/63714217/how-can-i-extract-numbers-containing-commas-from
                # -strings-in-python
                numbers = re.findall(r'\b\d{1,3}(?:,\d{3})*(?:\.\d+)?(?!\d)', benefit_item)
                if numbers:  # if we found salary data, return it
                    return int(float(numbers[0].replace(',', ''))), int(float(numbers[1].replace(',', '')))
            numbers = re.findall(r'\b\d{1,3}(?:,\d{3})*(?:\.\d+)?(?!\d)', benefit_item)
            if len(numbers) == 2 and int(
                    float(numbers[0].replace(',', ''))) > 30:  # some jobs just put the numbers in one item
                # and the description in another
                return int(float(numbers[0].replace(',', ''))), int(float(numbers[1].replace(',', '')))
            else:
                return min_salary, max_salary
    location = job_description.find("salary range")
    if location < 0:
        location = job_description.find("pay range")
    if location < 0:
        return min_salary, max_salary
    numbers = re.findall(r'\b\d{1,3}(?:,\d{3})*(?:\.\d+)?(?!\d)', job_description[location:location + 50])
    if numbers:
        return int(numbers[0].replace(',', '')), int(numbers[1].replace(',', ''))
    return min_salary, max_salary


def display_data(data: list):
    qt_app = PySide6.QtWidgets.QApplication(sys.argv)  # sys.argv is the list of command line arguments
    my_window = FirstWindow(data)
    assert my_window is not None
    sys.exit(qt_app.exec())


def start_gui(data):
    QQuickWindow.setGraphicsApi(QSGRendererInterface.GraphicsApi.Software)
    display_data(data)


def main():
    connection, cursor = open_database("job_search.sqlite")
    setup_database(cursor)
    pages_searched = 0
    total_pages = 5
    while pages_searched < total_pages:
        jobs = job_search(pages_searched * 10)
        jobs_data = get_jobs_data(jobs)
        insert_jobs_to_database(jobs_data, cursor)
        pages_searched += 1
    excel_jobs = get_excel_jobs()
    insert_jobs_to_database(excel_jobs, cursor)
    cursor.execute('''SELECT * from jobs;''')
    data = cursor.fetchall()
    close_database(connection)
    start_gui(data)


if __name__ == "__main__":
    main()
